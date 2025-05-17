# PythonEngine/main.py

import sys
import os
from file_parser import extract_content, extract_from_image
from analysis_prompt import run_analysis
from output_writer import save_output, save_as_word, save_as_txt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

def get_available_filename(base_path: str, extension: str) -> str:
    counter = 1
    new_path = f"{base_path}.{extension}"
    while os.path.exists(new_path):
        new_path = f"{base_path} ({counter}).{extension}"
        counter += 1
    return new_path

def parse_and_export_excel(output_txt_path: str, output_excel_path: str):
    with open(output_txt_path, "r", encoding="utf-8") as file:
        content = file.read()

    # Ambil baris yang merupakan bagian dari tabel markdown
    lines = [line for line in content.splitlines() if '|' in line]
    if len(lines) < 2:
        print("[INFO] Tidak ada tabel markdown untuk diekspor ke Excel.")
        return

    headers = [h.strip() for h in lines[0].strip('|').split('|')]
    rows = [line.strip('|').split('|') for line in lines[2:]]  # Skip garis pemisah

    wb = Workbook()
    ws = wb.active
    ws.append(headers)

    # Regex untuk mendeteksi formula Excel dalam teks
    formula_regex = r'=\s*([A-Z]+[0-9]+(?:\s*[\+\-\*\/]\s*[A-Z]+[0-9]+)*)'
    sum_regex = r'SUM\s*\(\s*([A-Z]+[0-9]+)\s*:\s*([A-Z]+[0-9]+)\s*\)'
    average_regex = r'AVERAGE\s*\(\s*([A-Z]+[0-9]+)\s*:\s*([A-Z]+[0-9]+)\s*\)'
    
    for row in rows:
        cells = [cell.strip() for cell in row]
        row_data = []
        
        for cell in cells:
            # Cek apakah cell berisi formula Excel
            if cell.startswith('='):
                # Pastikan formula ditulis dengan benar dan tambahkan langsung
                row_data.append(cell)
            # Cek kata kunci formula yang umum dan konversi jika perlu
            elif 'SUM(' in cell.upper() or 'JUMLAH(' in cell.upper():
                # Coba ekstrak range dan buat formula SUM yang benar
                sum_match = re.search(sum_regex, cell.upper())
                if sum_match:
                    row_data.append(f"=SUM({sum_match.group(1)}:{sum_match.group(2)})")
                else:
                    # Jika tidak match dengan format standar, tambahkan sebagai teks
                    row_data.append(cell)
            elif 'AVERAGE(' in cell.upper() or 'RATA-RATA(' in cell.upper():
                # Coba ekstrak range dan buat formula AVERAGE yang benar
                avg_match = re.search(average_regex, cell.upper())
                if avg_match:
                    row_data.append(f"=AVERAGE({avg_match.group(1)}:{avg_match.group(2)})")
                else:
                    row_data.append(cell)
            else:
                # Cek apakah ada operasi matematika sederhana yang mungkin formula
                formula_match = re.search(formula_regex, cell)
                if formula_match:
                    row_data.append(f"={formula_match.group(1)}")
                else:
                    row_data.append(cell)
        
        ws.append(row_data)

    # Auto-fit lebar kolom
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_excel_path)
    print(f"[SUKSES] File Excel hasil perbaikan disimpan ke {output_excel_path}")

def main():
    if len(sys.argv) < 6:
        print("Usage: python main.py <file_path or 'none'> <output_txt_path> <user_prompt> <output_format> <mode>")
        return

    file_path = sys.argv[1]
    output_txt_path = sys.argv[2]
    user_prompt = sys.argv[3]
    output_format = sys.argv[4].lower()  # txt, excel, word
    mode = sys.argv[5].lower()           # file, ocr, prompt-only

    if mode == "file":
        content = extract_content(file_path)
    elif mode == "ocr":
        content = extract_from_image(file_path)
    elif mode == "prompt-only":
        content = ""
    else:
        print(f"[ERROR] Mode tidak dikenali: {mode}")
        return

    result_text = run_analysis(content, user_prompt, output_format)
    save_output(result_text, output_txt_path)

    # Tentukan path aman untuk output
    base_path = output_txt_path.rsplit(".", 1)[0]

    if output_format == "txt":
        final_txt_path = get_available_filename(base_path + "_final", "txt")
        save_as_txt(result_text, final_txt_path)

    elif output_format == "excel":
        safe_excel_path = get_available_filename(base_path + "_parsed", "xlsx")
        parse_and_export_excel(output_txt_path, safe_excel_path)

    elif output_format == "word":
        safe_docx_path = get_available_filename(base_path + "_output", "docx")
        save_as_word(result_text, safe_docx_path)

    else:
        print(f"[ERROR] Format output tidak dikenali: {output_format}")
        return

    print("OK")

if __name__ == "__main__":
    main()
